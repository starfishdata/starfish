# Copyright (c) Meta Platforms, Inc. and affiliates.
# All rights reserved.
#
# This source code is licensed under the terms described in the LICENSE file in
# the root directory of this source tree.
# Excel parser logic

import os
from typing import Dict, Any
from starfish.data_ingest.parsers.base_parser import BaseParser


class ExcelParser(BaseParser):
    """Parser for Excel files"""

    def __init__(self):
        super().__init__()
        self.supported_extensions = [".xlsx", ".xls"]
        self.metadata = {}

    def parse(self, file_path: str) -> str:
        """Parse an Excel file into text

        Args:
            file_path: Path to the Excel file

        Returns:
            Extracted text from the Excel file
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel parsing. Install it with: pip install openpyxl")

        # Load workbook and extract metadata
        wb = openpyxl.load_workbook(file_path)
        self.metadata = {
            "file_path": file_path,
            "sheets": wb.sheetnames,
            "creator": wb.properties.creator,
            "created": wb.properties.created,
            "modified": wb.properties.modified,
            "last_modified_by": wb.properties.lastModifiedBy,
        }

        # Extract text from all sheets
        all_text = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_text = []
            sheet_text.append(f"--- Sheet: {sheet_name} ---")

            for row in sheet.iter_rows(values_only=True):
                row_text = [str(cell) for cell in row if cell is not None]
                if row_text:
                    sheet_text.append("\t".join(row_text))

            all_text.append("\n".join(sheet_text))

        return "\n\n".join(all_text)

    def get_metadata(self) -> Dict[str, Any]:
        """Get Excel file metadata

        Returns:
            Dictionary containing file metadata
        """
        return self.metadata

    def is_supported(self, file_path: str) -> bool:
        """Check if the file is supported by this parser

        Args:
            file_path: Path to the file

        Returns:
            True if the file is supported, False otherwise
        """
        return os.path.splitext(file_path)[1].lower() in self.supported_extensions
